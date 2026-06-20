from common import *
import csv


class ExcelExportMixin:
    def _validate_marker_sizes_for_export(self):
        invalid = []
        for i, m in enumerate(self.markers):
            try:
                size = float(m.get('size'))
                if not math.isfinite(size) or size <= 0:
                    raise ValueError
            except (TypeError, ValueError):
                invalid.append(m.get('name', f"Marker-{i + 1}"))
        if invalid:
            messagebox.showwarning(
                T("warn_title"),
                "Marker sizes must be positive numbers: " + ", ".join(invalid)
            )
            return False
        return True

    def export_to_excel(self):
        if not self.markers:
            messagebox.showwarning(T("warn_title"), T("warn_no_markers"))
            return
        if not self._validate_marker_sizes_for_export():
            return
        path = filedialog.asksaveasfilename(
            title=T("dlg_save_excel"),
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if not path:
            return
        try:
            import openpyxl
            from openpyxl.chart import ScatterChart, Reference, Series
            from openpyxl.chart.shapes import GraphicalProperties
            from openpyxl.drawing.line import LineProperties
            from openpyxl.chart.layout import Layout, ManualLayout

            def get_translated_roi_name(name):
                if name in ("MW Markers", "分子量マーカー", T('marker_node')):
                    return T('marker_node')
                return name

            wb = openpyxl.Workbook()
            ws1 = wb.active
            ws1.title = T('xl_sheet_cal')
            size_header = T('xl_size_kda') if self.mode == "protein" else T('xl_size_bp')
            log_header = T('xl_log') if self.mode == "protein" else T('xl_log_size')
            ws1.append([T('xl_marker_name'), T('xl_rf'), size_header, log_header])
            for m in self.markers:
                ws1.append([m['name'], m['rf'], m['size'], float(math.log10(m['size']))])

            ws2 = wb.create_sheet(title=T('xl_sheet_res'))
            ws2.append([T('xl_sample_no'), T('xl_sample_name'), T('xl_rf'), size_header])
            for i, s in enumerate(self.samples, 1):
                ws2.append([i, s['name'], s['rf'], s['size'] if s['size'] > 0 else "N/A"])

            if getattr(self, 'densitometry_rois', None):
                self._recalculate_densitometry()
                ws_den = wb.create_sheet(title=T('xl_sheet_dens'))
                ws_den.append([
                    "No.", "Lane/Sample", "X1", "Y1", "X2", "Y2",
                    "Integrated Density", "Relative to Max"
                ])
                for i, roi in enumerate(self.densitometry_rois, 1):
                    x0, y0, x1, y1 = roi.get('roi', [0, 0, 0, 0])
                    ws_den.append([
                        i,
                        get_translated_roi_name(roi.get('name', '')),
                        float(x0), float(y0), float(x1), float(y1),
                        float(roi.get('integrated_density', 0.0)),
                        float(roi.get('relative_density', 0.0)),
                    ])
                profile_start_col = 10
                max_len = 0
                for idx, roi in enumerate(self.densitometry_rois):
                    result = self._calculate_densitometry_profile(roi)
                    if not result:
                        continue
                    corrected = result.get('corrected', [])
                    max_len = max(max_len, len(corrected))
                    col_x = profile_start_col + idx * 2
                    col_y = col_x + 1
                    roi_name = get_translated_roi_name(roi.get('name', ''))
                    ws_den.cell(row=1, column=col_x, value=f"{roi_name} X")
                    ws_den.cell(row=1, column=col_y, value=f"{roi_name} Density")
                    xs = self._normalized_profile_x(len(corrected)) if hasattr(self, '_normalized_profile_x') else [
                        j / max(len(corrected) - 1, 1) for j in range(len(corrected))]
                    for row_idx, (xv, yv) in enumerate(zip(xs, corrected), start=2):
                        ws_den.cell(row=row_idx, column=col_x, value=float(xv))
                        ws_den.cell(row=row_idx, column=col_y, value=float(yv))
                if max_len:
                    dens_chart = ScatterChart()
                    dens_chart.title = T("lane_profile_title")
                    dens_chart.title.overlay = False
                    
                    # 凡例を重ねて表示
                    dens_chart.legend.position = 'tr'
                    dens_chart.legend.overlay = True
                    
                    # プロットエリアを左下から右上に少し小さく配置（余白を確保して被りを防止）
                    dens_chart.layout = Layout(
                        manualLayout=ManualLayout(
                            xMode="edge",
                            yMode="edge",
                            x=0.05,
                            y=0.00,
                            w=0.90,
                            h=0.85,
                        )
                    )
                    
                    # グラフの枠線を非表示にする
                    dens_chart.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))

                    dens_chart.x_axis.title = T("lane_profile_x")
                    dens_chart.y_axis.title = T("lane_profile_y")
                    dens_chart.x_axis.scaling.min = 0.0
                    dens_chart.x_axis.scaling.max = 1.0

                    # 軸の設定（目盛り内向き、目盛り線非表示、縦軸・横軸の追加、軸ラベルはlowに）
                    for axis in (dens_chart.x_axis, dens_chart.y_axis):
                        axis.graphicalProperties = GraphicalProperties(ln=LineProperties(solidFill="000000", w=12700))
                        axis.delete = False
                        axis.tickLblPos = "low"
                        axis.majorTickMark = "in"
                        axis.minorTickMark = "none"
                        axis.majorGridlines = None

                    for idx, roi in enumerate(self.densitometry_rois):
                        col_x = profile_start_col + idx * 2
                        col_y = col_x + 1
                        if ws_den.cell(row=2, column=col_y).value is None:
                            continue
                        xvalues = Reference(ws_den, min_col=col_x, min_row=2, max_row=max_len + 1)
                        yvalues = Reference(ws_den, min_col=col_y, min_row=2, max_row=max_len + 1)
                        roi_name = get_translated_roi_name(roi.get('name', ''))
                        series = Series(yvalues, xvalues, title=roi_name)
                        series.marker.symbol = "none"
                        
                        # アプリの線の色と同期
                        roi_color = self._get_densitometry_color(roi)
                        rgb_hex = roi_color.lstrip('#') if roi_color.startswith('#') else roi_color
                        series.graphicalProperties.line.solidFill = rgb_hex
                        series.graphicalProperties.line.width = 19050  # 1.5pt

                        dens_chart.series.append(series)
                    dens_chart.width = 16.18
                    dens_chart.height = 10.0
                    ws_den.add_chart(dens_chart, "J8")

            ws3 = wb.create_sheet(title=T('xl_sheet_graph'))

            # ---- 回帰直線用データを ws3 に書き込む ----
            # 列レイアウト: A=Rf(マーカー), B=log(Size)(マーカー), C=Rf(直線), D=log(Size)(直線)
            ws3.cell(row=1, column=1, value="Rf (Marker)")
            ws3.cell(row=1, column=2, value="log Size (Marker)")
            ws3.cell(row=1, column=3, value="Rf (Fit)")
            ws3.cell(row=1, column=4, value="log Size (Fit)")

            # マーカーデータを書き込む
            for i, m in enumerate(self.markers, start=2):
                ws3.cell(row=i, column=1, value=float(m['rf']))
                ws3.cell(row=i, column=2, value=float(math.log10(m['size'])))

            # 回帰直線データ（21点: 0.0〜1.0）を書き込む
            n_line_pts = 21
            for j in range(n_line_pts):
                rf_val = j / (n_line_pts - 1)
                log_val = float(self.calibration_a * rf_val + self.calibration_b)
                ws3.cell(row=j + 2, column=3, value=round(rf_val, 4))
                ws3.cell(row=j + 2, column=4, value=round(log_val, 6))

            # ---- ScatterChart を作成 ----
            chart = ScatterChart()
            chart.title = T('xl_cal_curve')
            chart.title.overlay = False
            
            # 凡例を重ねて表示
            chart.legend.position = 'tr'
            chart.legend.overlay = True

            # プロットエリアを左下から右上に少し小さく配置（余白を確保して被りを防止）
            chart.layout = Layout(
                manualLayout=ManualLayout(
                    xMode="edge",
                    yMode="edge",
                    x=0.15,
                    y=0.10,
                    w=0.70,
                    h=0.70
                )
            )

            # style=2: 白背景・テーマカラー非依存のシンプルスタイル
            # chart.style = 2

            # グラフの枠線を非表示にする
            chart.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))

            chart.x_axis.title = T('xl_xlabel')
            chart.y_axis.title = (T('xl_ylabel_kda') if self.mode == "protein"
                                  else T('xl_ylabel_bp'))
            chart.x_axis.numFmt = '0.00'
            chart.y_axis.numFmt = '0.00'
            chart.x_axis.scaling.min = 0.0
            chart.x_axis.scaling.max = 1.05

            # ---- 軸: 目盛りあり（内向き）・グリッド線なし・軸線の追加 ----
            for axis in (chart.x_axis, chart.y_axis):
                axis.graphicalProperties = GraphicalProperties(ln=LineProperties(solidFill="000000", w=12700))
                axis.delete = False
                axis.tickLblPos = "low"       # 軸ラベルを軸の外側に配置
                axis.majorTickMark = "in"     # 目盛りを内側に表示
                axis.minorTickMark = "none"
                axis.majorGridlines = None    # 主目盛り線（グリッド）を非表示

            # 凡例を右上に配置（回帰直線と被らない）
            chart.legend.position = 'tr'

            n_markers = len(self.markers)

            # ================================================================
            # Series 1: マーカーデータ（黒丸・塗りつぶし・接続線なし）
            # ================================================================
            xvalues_m = Reference(ws3, min_col=1, min_row=2, max_row=n_markers + 1)
            yvalues_m = Reference(ws3, min_col=2, min_row=2, max_row=n_markers + 1)
            series_m = Series(yvalues_m, xvalues_m, title="Marker")

            # マーカー形状: 塗りつぶし黒丸・サイズ7pt
            series_m.marker.symbol = "circle"
            series_m.marker.size = 7
            series_m.marker.graphicalProperties.solidFill = "000000"           # 塗り: 黒
            series_m.marker.graphicalProperties.line.solidFill = "000000"      # 輪郭: 黒
            series_m.marker.graphicalProperties.line.width = 9525              # 輪郭線太さ: 0.75pt

            # 各データ点を結ぶ線: なし
            series_m.graphicalProperties.line.noFill = True

            chart.series.append(series_m)

            # ================================================================
            # Series 2: 回帰直線（黒・1.5pt・マーカーなし）
            # ================================================================
            fit_label = (f"y={self.calibration_a:.3f}x+{self.calibration_b:.3f}"
                         f"  R\u00b2={self.calibration_r2:.4f}")
            xvalues_l = Reference(ws3, min_col=3, min_row=2, max_row=n_line_pts + 1)
            yvalues_l = Reference(ws3, min_col=4, min_row=2, max_row=n_line_pts + 1)
            series_l = Series(yvalues_l, xvalues_l, title=fit_label)

            # 点なし
            series_l.marker.symbol = "none"

            # 線: 黒・1.5pt (1pt = 12700 EMU)
            series_l.graphicalProperties.line.solidFill = "000000"
            series_l.graphicalProperties.line.width = 19050   # 1.5pt

            chart.series.append(series_l)

            # ================================================================
            # グラフサイズ
            # ================================================================
            chart.width = 16.18
            chart.height = 10.0

            ws3.add_chart(chart, "A1")

            wb.save(path)
            messagebox.showinfo(T("ok_title"), T("ok_excel"))
        except Exception as e:
            messagebox.showerror(T("err_title"), T("err_excel").format(e=e))

    def export_to_csv(self):
        """Export marker and sample data to CSV file."""
        if not self.markers:
            messagebox.showwarning(T("warn_title"), T("warn_no_markers"))
            return
        if not self._validate_marker_sizes_for_export():
            return
        path = filedialog.asksaveasfilename(
            title=T("dlg_save_csv"),
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv")]
        )
        if not path:
            return
        try:
            with open(path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                size_header = T('xl_size_kda') if self.mode == "protein" else T('xl_size_bp')
                log_header = T('xl_log') if self.mode == "protein" else T('xl_log_size')
                # Marker data
                writer.writerow([T('xl_marker_name'), T('xl_rf'), size_header, log_header])
                for m in self.markers:
                    writer.writerow([m['name'], m['rf'], m['size'], float(math.log10(m['size']))])
                writer.writerow([])
                # Sample data
                writer.writerow([T('xl_sample_no'), T('xl_sample_name'), T('xl_rf'), size_header])
                for i, s in enumerate(self.samples, 1):
                    writer.writerow([i, s['name'], s['rf'], s['size'] if s['size'] > 0 else "N/A"]) 
            messagebox.showinfo(T("ok_title"), T("ok_csv"))
        except Exception as e:
            messagebox.showerror(T("err_title"), T("err_csv").format(e=e))



    # ------------------------------------------------------------------ #
    #  アノテーション画像出力
    # ------------------------------------------------------------------ #

